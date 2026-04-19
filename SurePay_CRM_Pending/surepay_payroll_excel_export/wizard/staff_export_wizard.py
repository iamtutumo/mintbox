from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
from datetime import datetime
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("OpenPyXL not available, Excel export will not work")

class StaffExportWizard(models.TransientModel):
    _name = 'staff.export.wizard'
    _description = 'Staff Export Wizard for Payroll Processing'

    template_id = fields.Many2one('staff.export.template', string='Export Template', required=True)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    
    # Employee selection
    employee_ids = fields.Many2many('hr.employee', string='Employees')
    department_id = fields.Many2one('hr.department', string='Department')
    
    # Date range
    date_from = fields.Date(string='Date From', default=fields.Date.today)
    date_to = fields.Date(string='Date To', default=fields.Date.today)
    
    # Export options
    include_current_data = fields.Boolean(string='Include Current Payroll Data', default=True)
    include_empty_editable_fields = fields.Boolean(string='Include Empty Editable Fields', default=True)
    include_instructions = fields.Boolean(string='Include Instructions Sheet', default=True)
    include_validation = fields.Boolean(string='Include Validation Rules', default=True)
    
    # File options
    file_name = fields.Char(string='File Name', default='Staff_Payroll_Data')
    
    @api.model
    def default_get(self, fields):
        res = super().default_get(fields)
        
        # Get template from context
        template_id = self.env.context.get('default_template_id')
        if template_id:
            template = self.env['staff.export.template'].browse(template_id)
            if template:
                res['template_id'] = template.id
                res['file_name'] = f"{template.file_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        # Get selected employees from context
        if self.env.context.get('active_model') == 'hr.employee':
            employee_ids = self.env.context.get('active_ids', [])
            if employee_ids:
                res['employee_ids'] = [(6, 0, employee_ids)]
        
        return res
    
    def action_export_staff_data(self):
        """Export staff data to Excel"""
        self.ensure_one()
        
        if not self.employee_ids and not self.department_id:
            raise UserError(_("Please select employees or a department to export"))
        
        # Get employees to export
        employees = self._get_employees_to_export()
        
        if not employees:
            raise UserError(_("No employees found for the selected criteria"))
        
        # Generate Excel file
        excel_data = self._generate_excel_file(employees)
        
        # Create attachment
        filename = f"{self.file_name}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(excel_data),
            'res_model': 'staff.export.wizard',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def _get_employees_to_export(self):
        """Get employees to export based on selection criteria"""
        domain = []
        
        if self.employee_ids:
            domain.append(('id', 'in', self.employee_ids.ids))
        
        if self.department_id:
            domain.append(('department_id', '=', self.department_id.id))
        
        employees = self.env['hr.employee'].search(domain)
        return employees
    
    def _generate_excel_file(self, employees):
        """Generate Excel file with staff data"""
        try:
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create main data sheet
            data_sheet = workbook.create_sheet("Staff Payroll Data", 0)
            self._create_data_sheet(data_sheet, employees)
            
            # Create instructions sheet if requested
            if self.include_instructions:
                instructions_sheet = workbook.create_sheet("Instructions", 1)
                self._create_instructions_sheet(instructions_sheet)
            
            # Create validation sheet if requested
            if self.include_validation:
                validation_sheet = workbook.create_sheet("Validation Rules", 2)
                self._create_validation_sheet(validation_sheet)
            
            # Save to bytes
            excel_bytes = io.BytesIO()
            workbook.save(excel_bytes)
            excel_bytes.seek(0)
            
            return excel_bytes.read()
            
        except Exception as e:
            _logger.error(f"Error generating Excel file: {str(e)}")
            raise UserError(_("Error generating Excel file: %s") % str(e))
    
    def _create_data_sheet(self, sheet, employees):
        """Create the main data sheet"""
        template = self.template_id
        
        # Define headers
        headers = []
        editable_cols = []
        
        if template.include_basic_info:
            headers.extend(['Employee Code', 'Employee Name', 'Department', 'Position'])
        
        if template.include_current_salary and self.include_current_data:
            headers.extend(['Current Basic Salary', 'Current House Allowance', 'Current Transport Allowance', 
                           'Current Meal Allowance', 'Current Other Allowances'])
        
        # Editable salary fields
        if template.editable_basic_salary:
            headers.append('New Basic Salary')
            editable_cols.append(len(headers) - 1)
        
        if template.editable_allowances:
            headers.extend(['New House Allowance', 'New Transport Allowance', 'New Meal Allowance', 'New Other Allowances'])
            editable_cols.extend(range(len(headers) - 4, len(headers)))
        
        if template.editable_fines:
            headers.append('New Fine Deduction')
            editable_cols.append(len(headers) - 1)
        
        if template.editable_overtime:
            headers.extend(['Overtime Hours', 'Overtime Rate'])
            editable_cols.extend(range(len(headers) - 2, len(headers)))
        
        if template.editable_bonus:
            headers.append('Bonus Amount')
            editable_cols.append(len(headers) - 1)
        
        if template.editable_commission:
            headers.append('Commission Amount')
            editable_cols.append(len(headers) - 1)
        
        headers.extend(['Working Days', 'Leave Days', 'Sick Days'])
        editable_cols.extend(range(len(headers) - 3, len(headers)))
        
        # Style definitions
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=template.header_color, end_color=template.header_color, fill_type='solid')
        editable_fill = PatternFill(start_color=template.editable_color, end_color=template.editable_color, fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
        
        # Write employee data
        for row_num, employee in enumerate(employees, 2):
            col_num = 1
            
            # Basic info
            if template.include_basic_info:
                sheet.cell(row=row_num, column=col_num, value=employee.code or '').border = border
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=employee.name or '').border = border
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=employee.department_id.name or '').border = border
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=employee.job_id.name or '').border = border
                col_num += 1
            
            # Current salary data
            if template.include_current_salary and self.include_current_data:
                # Get current salary data from employee contract
                contract = employee.contract_id
                basic_salary = contract.wage if contract else 0
                
                sheet.cell(row=row_num, column=col_num, value=basic_salary).border = border
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=0).border = border  # Placeholder
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=0).border = border  # Placeholder
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=0).border = border  # Placeholder
                col_num += 1
                sheet.cell(row=row_num, column=col_num, value=0).border = border  # Placeholder
                col_num += 1
            
            # Editable fields (leave empty or with default values)
            editable_start_col = col_num
            for col in range(editable_start_col, len(headers) + 1):
                cell = sheet.cell(row=row_num, column=col, value='')
                cell.border = border
                
                # Highlight editable columns
                if (col - 1) in editable_cols:
                    cell.fill = editable_fill
                    
                    # Add default values for some fields
                    if headers[col - 1] == 'Working Days':
                        cell.value = template.default_working_days
                    elif headers[col - 1] == 'Overtime Rate':
                        cell.value = template.default_overtime_rate
        
        # Auto-adjust column widths
        for col_num in range(1, len(headers) + 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(headers[col_num - 1])), 15)
            sheet.column_dimensions[column_letter].width = max_length + 2
    
    def _create_instructions_sheet(self, sheet):
        """Create instructions sheet"""
        instructions = [
            ["Staff Payroll Data Export - Instructions"],
            [""],
            ["HOW TO USE THIS FILE:"],
            ["1. This file contains current employee payroll data and editable fields for adjustments"],
            ["2. Fill in the YELLOW highlighted cells with the new payroll data"],
            ["3. Do not modify the WHITE cells as they contain system data"],
            ["4. Save the file after making all necessary changes"],
            ["5. Upload the modified file using the 'Import Staff Data' wizard"],
            [""],
            ["FIELD DESCRIPTIONS:"],
            ["Employee Code - Unique employee identifier"],
            ["Employee Name - Full name of the employee"],
            ["Department - Employee's department"],
            ["Position - Employee's job position"],
            ["Current Basic Salary - Employee's current basic salary (read-only)"],
            ["New Basic Salary - Updated basic salary (editable)"],
            ["New House Allowance - Updated house allowance amount (editable)"],
            ["New Transport Allowance - Updated transport allowance amount (editable)"],
            ["New Meal Allowance - Updated meal allowance amount (editable)"],
            ["New Other Allowances - Updated other allowances amount (editable)"],
            ["Overtime Hours - Number of overtime hours worked (editable)"],
            ["Overtime Rate - Overtime multiplier (editable)"],
            ["Bonus Amount - Any bonus amount to be paid (editable)"],
            ["Commission Amount - Any commission amount to be paid (editable)"],
            ["Working Days - Number of working days in the period (editable)"],
            ["Leave Days - Number of leave days taken (editable)"],
            ["Sick Days - Number of sick days taken (editable)"],
            [""],
            ["VALIDATION:"],
            ["- All numeric fields must contain valid numbers"],
            ["- Working days must be between 0 and 31"],
            ["- Overtime rate should typically be between 1.0 and 3.0"],
            ["- Salary changes greater than 50% will trigger warnings"],
            [""],
            ["SUPPORT:"],
            ["Contact HR department for assistance with payroll data processing"],
        ]
        
        # Style definitions
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row_num, row_data in enumerate(instructions, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
                
                if row_num == 1:  # Title
                    cell.font = title_font
                elif row_num in [3, 10, 25, 29]:  # Section headers
                    cell.font = header_font
        
        # Auto-adjust column widths
        sheet.column_dimensions['A'].width = 50
    
    def _create_validation_sheet(self, sheet):
        """Create validation rules sheet"""
        validation_rules = [
            ["Field", "Data Type", "Validation Rule", "Error Message"],
            ["New Basic Salary", "Number", ">= 0", "Basic salary cannot be negative"],
            ["New House Allowance", "Number", ">= 0", "House allowance cannot be negative"],
            ["New Transport Allowance", "Number", ">= 0", "Transport allowance cannot be negative"],
            ["New Meal Allowance", "Number", ">= 0", "Meal allowance cannot be negative"],
            ["New Other Allowances", "Number", ">= 0", "Other allowances cannot be negative"],
            ["Overtime Hours", "Number", ">= 0", "Overtime hours cannot be negative"],
            ["Overtime Rate", "Number", ">= 1.0 and <= 3.0", "Overtime rate should be between 1.0 and 3.0"],
            ["Bonus Amount", "Number", ">= 0", "Bonus amount cannot be negative"],
            ["Commission Amount", "Number", ">= 0", "Commission amount cannot be negative"],
            ["Working Days", "Integer", ">= 0 and <= 31", "Working days must be between 0 and 31"],
            ["Leave Days", "Integer", ">= 0", "Leave days cannot be negative"],
            ["Sick Days", "Integer", ">= 0", "Sick days cannot be negative"],
        ]
        
        # Style definitions
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='#366092', end_color='#366092', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        alignment = Alignment(horizontal='center', vertical='center')
        
        # Write headers
        for col_num, header in enumerate(validation_rules[0], 1):
            cell = sheet.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment
        
        # Write validation rules
        for row_num, rule_data in enumerate(validation_rules[1:], 2):
            for col_num, cell_value in enumerate(rule_data, 1):
                cell = sheet.cell(row=row_num, column=col_num, value=cell_value)
                cell.border = border
                cell.alignment = alignment
        
        # Auto-adjust column widths
        for col_num in range(1, len(validation_rules[0]) + 1):
            column_letter = get_column_letter(col_num)
            max_length = max(len(str(rule[col_num - 1])) for rule in validation_rules)
            sheet.column_dimensions[column_letter].width = max_length + 2

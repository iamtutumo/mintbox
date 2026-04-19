from odoo import models, fields, api

class PayrollExportTemplate(models.Model):
    _name = 'payroll.export.template'
    _description = 'Payroll Export Template'
    _order = 'name'

    name = fields.Char(string='Template Name', required=True)
    code = fields.Char(string='Code', required=True, help="Unique code for the template")
    description = fields.Text(string='Description')
    
    template_type = fields.Selection([
        ('payslip', 'Individual Payslip'),
        ('payroll_register', 'Payroll Register'),
        ('department_summary', 'Department Summary'),
        ('employee_history', 'Employee History'),
        ('custom', 'Custom Template'),
    ], string='Template Type', required=True, default='payslip')
    
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    active = fields.Boolean(string='Active', default=True)
    
    # Template configuration
    include_employee_info = fields.Boolean(string='Include Employee Information', default=True)
    include_salary_details = fields.Boolean(string='Include Salary Details', default=True)
    include_deductions = fields.Boolean(string='Include Deductions', default=True)
    include_allowances = fields.Boolean(string='Include Allowances', default=True)
    include_taxes = fields.Boolean(string='Include Taxes', default=True)
    include_net_salary = fields.Boolean(string='Include Net Salary', default=True)
    
    # Custom fields configuration
    custom_fields_ids = fields.Many2many('ir.model.fields', string='Custom Fields',
                                         domain="[('model', 'in', ['surepay_payroll.hr.payslip', 'surepay_payroll.hr.payslip.line'])]")
    
    # Excel formatting
    header_color = fields.Char(string='Header Color', default='#366092', help="Hex color code for headers")
    header_font_size = fields.Integer(string='Header Font Size', default=12)
    data_font_size = fields.Integer(string='Data Font Size', default=10)
    
    # Export settings
    file_prefix = fields.Char(string='File Prefix', default='Payroll_Export')
    include_company_logo = fields.Boolean(string='Include Company Logo', default=True)
    include_summary_sheet = fields.Boolean(string='Include Summary Sheet', default=True)
    
    @api.constrains('code')
    def _check_code_unique(self):
        for template in self:
            existing = self.search([
                ('code', '=', template.code),
                ('id', '!=', template.id),
                ('company_id', '=', template.company_id.id)
            ])
            if existing:
                raise UserError(f"Template code '{template.code}' must be unique per company.")
    
    def action_export_with_template(self, payslips):
        """Export payslips using this template"""
        self.ensure_one()
        
        if not payslips:
            raise UserError("No payslips selected for export.")
        
        # Generate Excel data based on template
        excel_data = self._generate_excel_from_template(payslips)
        
        # Create attachment
        filename = f"{self.file_prefix}_{fields.Date.today().strftime('%Y%m%d')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': fields.base64.b64encode(excel_data),
            'res_model': 'payroll.export.template',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def _generate_excel_from_template(self, payslips):
        """Generate Excel data based on template configuration"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("Please install openpyxl library to use Excel export functionality")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Generate sheets based on template type
        if self.template_type == 'payslip':
            self._generate_payslip_sheets(wb, payslips)
        elif self.template_type == 'payroll_register':
            self._generate_payroll_register_sheet(wb, payslips)
        elif self.template_type == 'department_summary':
            self._generate_department_summary_sheet(wb, payslips)
        elif self.template_type == 'employee_history':
            self._generate_employee_history_sheet(wb, payslips)
        else:
            self._generate_custom_sheet(wb, payslips)
        
        # Add summary sheet if requested
        if self.include_summary_sheet:
            self._generate_summary_sheet(wb, payslips)
        
        # Save to bytes
        import io
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def _generate_payslip_sheets(self, wb, payslips):
        """Generate individual sheets for each payslip"""
        for payslip in payslips:
            ws = wb.create_sheet(title=f"{payslip.employee_id.name[:20]}")
            self._populate_payslip_sheet(ws, payslip)
    
    def _populate_payslip_sheet(self, ws, payslip):
        """Populate a single payslip sheet"""
        # Define styles
        header_font = Font(name='Arial', size=self.header_font_size, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(name='Arial', size=self.header_font_size + 2, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        label_font = Font(name='Arial', size=self.data_font_size, bold=True)
        value_font = Font(name='Arial', size=self.data_font_size)
        
        # Add title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"PAYSLIP - {payslip.company_id.name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Add period
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Period: {payslip.date_from.strftime('%d %B %Y')} to {payslip.date_to.strftime('%d %B %Y')}"
        ws['A2'].font = Font(name='Arial', size=self.data_font_size, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        row = 4
        if self.include_employee_info:
            # Employee information
            ws[f'A{row}'] = "Employee Information"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            employee_info = [
                ("Employee Name:", payslip.employee_id.name),
                ("Employee ID:", payslip.employee_id.work_email or ''),
                ("Department:", payslip.employee_id.department_id.name if payslip.employee_id.department_id else ''),
                ("Job Position:", payslip.employee_id.job_id.name if payslip.employee_id.job_id else ''),
            ]
            
            for label, value in employee_info:
                ws[f'A{row}'] = label
                ws[f'A{row}'].font = label_font
                ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
                
                ws[f'B{row}'] = value
                ws[f'B{row}'].font = value_font
                ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(f'B{row}:D{row}')
                row += 1
        
        if self.include_salary_details:
            # Salary details
            row += 1
            ws[f'A{row}'] = "Salary Details"
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].alignment = header_alignment
            ws.merge_cells(f'A{row}:D{row}')
            
            row += 1
            # Headers
            headers = ["Description", "Code", "Amount", "Notes"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            row += 1
            # Salary lines
            for line in payslip.line_ids.sorted('sequence'):
                if line.appears_on_payslip:
                    ws.cell(row=row, column=1, value=line.name)
                    ws.cell(row=row, column=2, value=line.code or '')
                    ws.cell(row=row, column=3, value=line.total)
                    ws.cell(row=row, column=4, value=line.note or '')
                    
                    # Apply formatting
                    for col in range(1, 5):
                        cell = ws.cell(row=row, column=col)
                        cell.font = value_font
                        cell.alignment = Alignment(horizontal='left' if col in [1, 2, 4] else 'right', vertical='center')
                    
                    row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            column_letter = get_column_letter(col)
            max_length = 0
            for row_obj in ws.iter_rows():
                for cell in row_obj:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _generate_payroll_register_sheet(self, wb, payslips):
        """Generate payroll register sheet"""
        ws = wb.create_sheet(title="Payroll Register")
        
        # Define styles
        header_font = Font(name='Arial', size=self.header_font_size, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=self.header_color, end_color=self.header_color, fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = ["Employee", "Department", "Period", "Gross Salary", "Deductions", "Net Salary", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        row = 2
        for payslip in payslips.sorted('employee_id.name'):
            ws.cell(row=row, column=1, value=payslip.employee_id.name)
            ws.cell(row=row, column=2, value=payslip.employee_id.department_id.name if payslip.employee_id.department_id else '')
            ws.cell(row=row, column=3, value=f"{payslip.date_from.strftime('%d/%m/%Y')} - {payslip.date_to.strftime('%d/%m/%Y')}")
            ws.cell(row=row, column=4, value=payslip.gross_wage)
            ws.cell(row=row, column=5, value=payslip.total_ded)
            ws.cell(row=row, column=6, value=payslip.net_wage)
            ws.cell(row=row, column=7, value=dict(payslip._fields['state'].selection).get(payslip.state, ''))
            row += 1
    
    def _generate_department_summary_sheet(self, wb, payslips):
        """Generate department summary sheet"""
        ws = wb.create_sheet(title="Department Summary")
        
        # Group by department
        department_data = {}
        for payslip in payslips:
            dept = payslip.employee_id.department_id.name or 'No Department'
            if dept not in department_data:
                department_data[dept] = {
                    'count': 0,
                    'total_gross': 0,
                    'total_deductions': 0,
                    'total_net': 0,
                }
            department_data[dept]['count'] += 1
            department_data[dept]['total_gross'] += payslip.gross_wage
            department_data[dept]['total_deductions'] += payslip.total_ded
            department_data[dept]['total_net'] += payslip.net_wage
        
        # Headers
        headers = ["Department", "Employee Count", "Total Gross", "Total Deductions", "Total Net"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        row = 2
        for dept, data in department_data.items():
            ws.cell(row=row, column=1, value=dept)
            ws.cell(row=row, column=2, value=data['count'])
            ws.cell(row=row, column=3, value=data['total_gross'])
            ws.cell(row=row, column=4, value=data['total_deductions'])
            ws.cell(row=row, column=5, value=data['total_net'])
            row += 1
    
    def _generate_employee_history_sheet(self, wb, payslips):
        """Generate employee history sheet"""
        ws = wb.create_sheet(title="Employee History")
        
        # Group by employee
        employee_data = {}
        for payslip in payslips:
            emp = payslip.employee_id.name
            if emp not in employee_data:
                employee_data[emp] = []
            employee_data[emp].append(payslip)
        
        # Headers
        headers = ["Employee", "Period", "Gross Salary", "Deductions", "Net Salary", "Status"]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        row = 2
        for emp, emp_payslips in employee_data.items():
            for payslip in sorted(emp_payslips, key=lambda p: p.date_from):
                ws.cell(row=row, column=1, value=emp)
                ws.cell(row=row, column=2, value=f"{payslip.date_from.strftime('%d/%m/%Y')} - {payslip.date_to.strftime('%d/%m/%Y')}")
                ws.cell(row=row, column=3, value=payslip.gross_wage)
                ws.cell(row=row, column=4, value=payslip.total_ded)
                ws.cell(row=row, column=5, value=payslip.net_wage)
                ws.cell(row=row, column=6, value=dict(payslip._fields['state'].selection).get(payslip.state, ''))
                row += 1
    
    def _generate_custom_sheet(self, wb, payslips):
        """Generate custom sheet based on template configuration"""
        ws = wb.create_sheet(title="Custom Export")
        
        # Basic headers
        headers = ["Employee", "Department", "Period"]
        if self.include_salary_details:
            headers.extend(["Gross Salary", "Basic Salary"])
        if self.include_deductions:
            headers.append("Total Deductions")
        if self.include_allowances:
            headers.append("Total Allowances")
        if self.include_taxes:
            headers.append("Total Taxes")
        if self.include_net_salary:
            headers.append("Net Salary")
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data rows
        row = 2
        for payslip in payslips:
            col = 1
            ws.cell(row=row, column=col, value=payslip.employee_id.name)
            col += 1
            ws.cell(row=row, column=col, value=payslip.employee_id.department_id.name if payslip.employee_id.department_id else '')
            col += 1
            ws.cell(row=row, column=col, value=f"{payslip.date_from.strftime('%d/%m/%Y')} - {payslip.date_to.strftime('%d/%m/%Y')}")
            col += 1
            
            if self.include_salary_details:
                ws.cell(row=row, column=col, value=payslip.gross_wage)
                col += 1
                # Get basic salary
                basic_salary = 0
                for line in payslip.line_ids:
                    if line.code == 'BASIC':
                        basic_salary = line.total
                        break
                ws.cell(row=row, column=col, value=basic_salary)
                col += 1
            
            if self.include_deductions:
                ws.cell(row=row, column=col, value=payslip.total_ded)
                col += 1
            
            if self.include_allowances:
                # Calculate total allowances
                allowances = 0
                for line in payslip.line_ids:
                    if line.category_id.code == 'ALW' and line.total > 0:
                        allowances += line.total
                ws.cell(row=row, column=col, value=allowances)
                col += 1
            
            if self.include_taxes:
                # Calculate total taxes
                taxes = 0
                for line in payslip.line_ids:
                    if line.category_id.code == 'TAX' and line.total < 0:
                        taxes += abs(line.total)
                ws.cell(row=row, column=col, value=taxes)
                col += 1
            
            if self.include_net_salary:
                ws.cell(row=row, column=col, value=payslip.net_wage)
            
            row += 1
    
    def _generate_summary_sheet(self, wb, payslips):
        """Generate summary sheet"""
        ws = wb.create_sheet(title="Summary")
        
        # Calculate totals
        total_employees = len(payslips.mapped('employee_id'))
        total_gross = sum(payslips.mapped('gross_wage'))
        total_deductions = sum(payslips.mapped('total_ded'))
        total_net = sum(payslips.mapped('net_wage'))
        
        # Summary data
        summary_data = [
            ["Total Employees", total_employees],
            ["Total Gross Salary", total_gross],
            ["Total Deductions", total_deductions],
            ["Total Net Salary", total_net],
            ["Average Gross Salary", total_gross / total_employees if total_employees > 0 else 0],
            ["Average Net Salary", total_net / total_employees if total_employees > 0 else 0],
        ]
        
        # Headers
        ws.cell(row=1, column=1, value="Summary")
        ws.cell(row=1, column=2, value="Amount")
        
        # Data
        row = 2
        for label, value in summary_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

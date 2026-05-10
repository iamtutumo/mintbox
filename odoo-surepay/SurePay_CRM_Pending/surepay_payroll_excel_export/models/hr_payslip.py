from odoo import models, fields, api
import base64
import io
from datetime import datetime

class HrPayslip(models.Model):
    _inherit = 'surepay_payroll.hr.payslip'

    def action_export_to_excel(self):
        """Export single payslip to Excel"""
        self.ensure_one()
        
        # Create Excel data
        excel_data = self._generate_payslip_excel()
        
        # Create attachment
        filename = f"Payslip_{self.employee_id.name}_{self.date_from.strftime('%Y%m%d')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(excel_data),
            'res_model': 'surepay_payroll.hr.payslip',
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def action_export_multiple_to_excel(self):
        """Export multiple payslips to Excel"""
        if len(self) == 1:
            return self.action_export_to_excel()
        
        # Create Excel data for multiple payslips
        excel_data = self._generate_payslips_excel()
        
        # Create attachment
        filename = f"Payslips_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(excel_data),
            'res_model': 'surepay_payroll.hr.payslip',
            'res_id': self[0].id if self else False,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
    
    def _generate_payslip_excel(self):
        """Generate Excel data for single payslip"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("Please install openpyxl library to use Excel export functionality")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Payslip - {self.employee_id.name}"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        label_font = Font(name='Arial', size=11, bold=True)
        label_alignment = Alignment(horizontal='left', vertical='center')
        
        value_font = Font(name='Arial', size=11)
        value_alignment = Alignment(horizontal='right', vertical='center')
        
        # Add title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"PAYSLIP - {self.company_id.name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Add period
        ws.merge_cells('A2:D2')
        ws['A2'] = f"Period: {self.date_from.strftime('%d %B %Y')} to {self.date_to.strftime('%d %B %Y')}"
        ws['A2'].font = Font(name='Arial', size=11, bold=True)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Employee information
        row = 4
        ws[f'A{row}'] = "Employee Information"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        employee_info = [
            ("Employee Name:", self.employee_id.name),
            ("Employee ID:", self.employee_id.work_email or ''),
            ("Department:", self.employee_id.department_id.name if self.employee_id.department_id else ''),
            ("Job Position:", self.employee_id.job_id.name if self.employee_id.job_id else ''),
        ]
        
        for label, value in employee_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].alignment = label_alignment
            
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = value_font
            ws[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Salary details
        row += 1
        ws[f'A{row}'] = "Salary Details"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws[f'A{row}'].alignment = header_alignment
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        # Headers
        headers = ["Description", "Code", "Amount", "Notes"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        row += 1
        # Salary lines
        for line in self.line_ids.sorted('sequence'):
            if line.appears_on_payslip:
                ws.cell(row=row, column=1, value=line.name)
                ws.cell(row=row, column=2, value=line.code or '')
                ws.cell(row=row, column=3, value=line.total)
                ws.cell(row=row, column=4, value=line.note or '')
                
                # Apply formatting
                for col in range(1, 5):
                    cell = ws.cell(row=row, column=col)
                    cell.font = value_font
                    cell.alignment = Alignment(horizontal='left' if col in [1, 2, 4] else 'right', vertical='center')
                
                row += 1
        
        # Totals
        row += 1
        totals = [
            ("Gross Salary:", self.gross_wage),
            ("Deductions:", self.total_ded),
            ("Net Salary:", self.net_wage),
        ]
        
        for label, amount in totals:
            ws[f'C{row}'] = label
            ws[f'C{row}'].font = label_font
            ws[f'C{row}'].alignment = label_alignment
            
            ws[f'D{row}'] = amount
            ws[f'D{row}'].font = Font(name='Arial', size=11, bold=True)
            ws[f'D{row}'].alignment = value_alignment
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 5):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()
    
    def _generate_payslips_excel(self):
        """Generate Excel data for multiple payslips"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise UserError("Please install openpyxl library to use Excel export functionality")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payslips Summary"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        title_font = Font(name='Arial', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        
        # Add title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"PAYROLL SUMMARY - {self.company_id.name}"
        ws['A1'].font = title_font
        ws['A1'].alignment = title_alignment
        
        # Add date
        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generated on: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
        ws['A2'].font = Font(name='Arial', size=11)
        ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        row = 4
        headers = ["Employee", "Department", "Period", "Gross Salary", "Deductions", "Net Salary", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data rows
        row += 1
        for payslip in self.sorted('employee_id.name'):
            ws.cell(row=row, column=1, value=payslip.employee_id.name)
            ws.cell(row=row, column=2, value=payslip.employee_id.department_id.name if payslip.employee_id.department_id else '')
            ws.cell(row=row, column=3, value=f"{payslip.date_from.strftime('%d/%m/%Y')} - {payslip.date_to.strftime('%d/%m/%Y')}")
            ws.cell(row=row, column=4, value=payslip.gross_wage)
            ws.cell(row=row, column=5, value=payslip.total_ded)
            ws.cell(row=row, column=6, value=payslip.net_wage)
            ws.cell(row=row, column=7, value=dict(payslip._fields['state'].selection).get(payslip.state, ''))
            
            # Apply formatting
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(name='Arial', size=10)
                cell.alignment = Alignment(horizontal='left' if col in [1, 2, 3, 7] else 'right', vertical='center')
            
            row += 1
        
        # Auto-adjust column widths
        for col in range(1, 8):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws.iter_rows():
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.read()

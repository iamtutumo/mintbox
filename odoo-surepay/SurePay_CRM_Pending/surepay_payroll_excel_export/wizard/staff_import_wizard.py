from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning("OpenPyXL not available, Excel import will not work")

class StaffImportWizard(models.TransientModel):
    _name = 'staff.import.wizard'
    _description = 'Staff Import Wizard for Payroll Processing'

    import_file = fields.Binary(string='Import File', required=True)
    import_filename = fields.Char(string='Filename')
    
    template_id = fields.Many2one('staff.export.template', string='Template', required=True)
    company_id = fields.Many2one('res.company', string='Company', default=lambda self: self.env.company)
    
    # Import options
    payroll_period_start = fields.Date(string='Payroll Period Start', required=True)
    payroll_period_end = fields.Date(string='Payroll Period End', required=True)
    send_payslips_email = fields.Boolean(string='Send Payslips by Email', default=True)
    email_template_id = fields.Many2one('mail.template', string='Email Template',
                                       domain="[('model', '=', 'surepay_payroll.hr.payslip')]")
    
    # Batch options
    batch_name = fields.Char(string='Batch Name', required=True)
    description = fields.Text(string='Description')
    
    # Processing options
    validate_only = fields.Boolean(string='Validate Only (Do Not Process)', default=False)
    skip_validation = fields.Boolean(string='Skip Validation (Not Recommended)', default=False)
    update_existing_payslips = fields.Boolean(string='Update Existing Payslips', default=True)
    
    # Preview data
    preview_data = fields.Text(string='Preview Data', readonly=True)
    has_errors = fields.Boolean(string='Has Errors', readonly=True)
    error_count = fields.Integer(string='Error Count', readonly=True)
    warning_count = fields.Integer(string='Warning Count', readonly=True)
    
    @api.model
    def default_get(self, fields):
        res = super().default_get(fields)
        
        # Get template from context
        template_id = self.env.context.get('default_template_id')
        if template_id:
            template = self.env['staff.export.template'].browse(template_id)
            if template:
                res['template_id'] = template.id
        
        # Generate default batch name
        res['batch_name'] = f"Staff Import {fields.Date.today().strftime('%Y-%m-%d')}"
        
        return res
    
    def action_validate_file(self):
        """Validate the uploaded Excel file"""
        self.ensure_one()
        
        try:
            # Read and validate Excel file
            excel_data = self._read_excel_file()
            validation_result = self._validate_excel_data(excel_data)
            
            # Update preview and error counts
            self.preview_data = self._format_preview_data(validation_result['data'][:10])  # Show first 10 rows
            self.has_errors = validation_result['error_count'] > 0
            self.error_count = validation_result['error_count']
            self.warning_count = validation_result['warning_count']
            
            if self.has_errors:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Validation Errors'),
                        'message': _('The file contains %d errors. Please fix them before importing.') % self.error_count,
                        'type': 'danger',
                    }
                }
            else:
                return {
                    'type': 'ir.actions.client',
                    'tag': 'display_notification',
                    'params': {
                        'title': _('Validation Successful'),
                        'message': _('The file is valid and ready for import. %d warnings found.') % self.warning_count,
                        'type': 'success',
                    }
                }
                
        except Exception as e:
            _logger.error(f"Error validating import file: {str(e)}")
            raise UserError(_("Error validating import file: %s") % str(e))
    
    def action_import_data(self):
        """Import the staff data and process payroll"""
        self.ensure_one()
        
        if not self.import_file:
            raise UserError(_("Please select a file to import"))
        
        if not self.payroll_period_start or not self.payroll_period_end:
            raise UserError(_("Please select both payroll period start and end dates"))
        
        if self.payroll_period_start > self.payroll_period_end:
            raise UserError(_("Payroll period start date must be before end date"))
        
        # Validate file first if not skipped
        if not self.skip_validation:
            self.action_validate_file()
            if self.has_errors:
                raise UserError(_("Please fix validation errors before importing"))
        
        try:
            # Create import batch
            batch_vals = {
                'name': self.batch_name,
                'description': self.description,
                'template_id': self.template_id.id,
                'import_file': self.import_file,
                'import_filename': self.import_filename,
                'payroll_period_start': self.payroll_period_start,
                'payroll_period_end': self.payroll_period_end,
                'send_payslips_email': self.send_payslips_email,
                'email_template_id': self.email_template_id.id if self.email_template_id else False,
            }
            
            batch = self.env['staff.import.batch'].create(batch_vals)
            
            # If validate only, return batch view
            if self.validate_only:
                return {
                    'type': 'ir.actions.act_window',
                    'name': 'Import Batch',
                    'res_model': 'staff.import.batch',
                    'res_id': batch.id,
                    'view_mode': 'form',
                    'target': 'current',
                }
            
            # Process the batch
            batch.action_validate_all()
            batch.action_process_batch()
            
            # Return batch view
            return {
                'type': 'ir.actions.act_window',
                'name': 'Import Batch',
                'res_model': 'staff.import.batch',
                'res_id': batch.id,
                'view_mode': 'form',
                'target': 'current',
            }
            
        except Exception as e:
            _logger.error(f"Error importing staff data: {str(e)}")
            raise UserError(_("Error importing staff data: %s") % str(e))
    
    def _read_excel_file(self):
        """Read Excel file and return data"""
        try:
            # Decode base64 file
            file_data = base64.b64decode(self.import_file)
            excel_file = io.BytesIO(file_data)
            
            # Load workbook
            workbook = openpyxl.load_workbook(excel_file, data_only=True)
            
            # Get main data sheet
            if 'Staff Payroll Data' in workbook.sheetnames:
                sheet = workbook['Staff Payroll Data']
            else:
                sheet = workbook.active
            
            # Read data (skip header row)
            data = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row[0]:  # Only process rows with employee code
                    data.append(row)
            
            return data
            
        except Exception as e:
            _logger.error(f"Error reading Excel file: {str(e)}")
            raise UserError(_("Error reading Excel file: %s") % str(e))
    
    def _validate_excel_data(self, excel_data):
        """Validate Excel data structure and content"""
        errors = []
        warnings = []
        valid_data = []
        
        for row_num, row in enumerate(excel_data, 2):  # Start from row 2 (after header)
            row_errors = []
            row_warnings = []
            
            # Check if row has enough columns
            if len(row) < 4:
                row_errors.append(f"Row {row_num}: Insufficient columns")
                continue
            
            # Extract data
            employee_code = str(row[0]) if row[0] else ''
            employee_name = str(row[1]) if row[1] else ''
            
            # Validate required fields
            if not employee_code:
                row_errors.append(f"Row {row_num}: Employee code is required")
            
            if not employee_name:
                row_errors.append(f"Row {row_num}: Employee name is required")
            
            # Find employee
            employee = self.env['hr.employee'].search([('code', '=', employee_code)], limit=1)
            if not employee:
                row_errors.append(f"Row {row_num}: Employee with code '{employee_code}' not found")
            
            # Validate numeric fields
            try:
                # Basic salary validation
                if len(row) > 5 and row[5]:  # New basic salary
                    basic_salary = float(row[5])
                    if basic_salary < 0:
                        row_errors.append(f"Row {row_num}: Basic salary cannot be negative")
                    
                    # Check for large salary changes
                    if employee and employee.contract_id:
                        current_salary = employee.contract_id.wage
                        if current_salary > 0:
                            change_percent = abs(basic_salary - current_salary) / current_salary * 100
                            if change_percent > 50:
                                row_warnings.append(f"Row {row_num}: Salary change of {change_percent:.1f}% is unusual")
                
                # Allowances validation
                for col_idx in range(6, min(10, len(row))):  # House, Transport, Meal, Other allowances
                    if row[col_idx]:
                        allowance = float(row[col_idx])
                        if allowance < 0:
                            row_errors.append(f"Row {row_num}: Allowance cannot be negative")
                
                # Overtime validation
                if len(row) > 10 and row[10]:  # Overtime hours
                    overtime_hours = float(row[10])
                    if overtime_hours < 0:
                        row_errors.append(f"Row {row_num}: Overtime hours cannot be negative")
                    elif overtime_hours > 200:
                        row_warnings.append(f"Row {row_num}: Overtime hours ({overtime_hours}) seem unusually high")
                
                if len(row) > 11 and row[11]:  # Overtime rate
                    overtime_rate = float(row[11])
                    if overtime_rate < 1.0 or overtime_rate > 3.0:
                        row_warnings.append(f"Row {row_num}: Overtime rate should be between 1.0 and 3.0")
                
                # Working days validation
                if len(row) > 14 and row[14]:  # Working days
                    working_days = int(row[14])
                    if working_days < 0 or working_days > 31:
                        row_errors.append(f"Row {row_num}: Working days must be between 0 and 31")
                
            except (ValueError, TypeError) as e:
                row_errors.append(f"Row {row_num}: Invalid numeric data - {str(e)}")
            
            # Add row to valid data if no errors
            if not row_errors:
                valid_data.append(row)
            
            # Add errors and warnings
            errors.extend(row_errors)
            warnings.extend(row_warnings)
        
        return {
            'data': valid_data,
            'errors': errors,
            'warnings': warnings,
            'error_count': len(errors),
            'warning_count': len(warnings),
        }
    
    def _format_preview_data(self, data):
        """Format preview data for display"""
        if not data:
            return "No data to preview"
        
        preview = "Preview of first 10 rows:\n\n"
        preview += "Row | Employee Code | Employee Name | Department | Status\n"
        preview += "-" * 70 + "\n"
        
        for row_num, row in enumerate(data[:10], 2):
            employee_code = str(row[0]) if row[0] else ''
            employee_name = str(row[1]) if row[1] else ''
            department = str(row[2]) if row[2] else ''
            
            # Find employee to get status
            employee = self.env['hr.employee'].search([('code', '=', employee_code)], limit=1)
            status = "Found" if employee else "Not Found"
            
            preview += f"{row_num:3d} | {employee_code:12s} | {employee_name:20s} | {department:15s} | {status}\n"
        
        if len(data) > 10:
            preview += f"\n... and {len(data) - 10} more rows"
        
        return preview
